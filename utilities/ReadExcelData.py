'''
Created on 02-Oct-2020

@author: abhinandan.sapkal
'''
import openpyxl
from openpyxl import cell

#Setting excel file path
filePath="../testDataFiles/TestData.xlsx"



def readData(filePath, sheetName, rowNumber, columnNumber):
    workbook = openpyxl.load_workbook(filePath)
    sheet= workbook.get_sheet_by_name("AddCandidate")
    cellValue = sheet.cell(row=rowNumber,column=columnNumber).value
    return cellValue
    
def totalRows(filePath, sheetName):
    workbook = openpyxl.load_workbook(filePath)
    sheet= workbook.get_sheet_by_name("AddCandidate")
    totalRowCount = sheet.max_row
    return totalRowCount
    
def totalColumns(filePath, sheetName):
    workbook = openpyxl.load_workbook(filePath)
    sheet= workbook.get_sheet_by_name("AddCandidate")
    totalColumnCount = sheet.max_column
    return totalColumnCount    
    
    
    
    
    
    